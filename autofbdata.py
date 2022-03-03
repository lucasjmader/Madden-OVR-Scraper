# INF360 - Programming in Python
# Lucas Mader
# Final

#setting up logging
import logging
logging.basicConfig(filename='finallog.txt', level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

#import sys to use the sys.exit() method
try:
    import sys
    logging.debug('sys module loaded successfully')
except:
    logging.critical('sys module not found')
    print('sys module not found')
    sys.exit()

#to install selenium: pip install selenium
try:
    from selenium import webdriver
    logging.debug('Selenium module loaded successfully')
except:
    logging.critical('Selenium module not found')
    logging.critical('Selenium is not installed. Please try running: pip install '
                     'selenium in the command prompt ')
    print('Selenium is not installed. Please try running pip install '
          'selenium in the command prompt')
    sys.exit()



#this imports the ability to control the keyboard
try:
    from selenium.webdriver.common.keys import Keys

    logging.debug('Keys loaded successfully')
except:
    logging.critical('Keys not found')
    print('Keys not found')
    sys.exit()

#the time module is built-in to python so it doesn't need installed

try:
    import time
    logging.debug('time loaded successfully')
except:
    logging.critical('time not found')
    print('time not found')
    sys.exit()


#import openpyxl
try:
    from openpyxl import Workbook, load_workbook
    logging.debug('openpyxl loaded successfully')
except:
    logging.critical(' not found')
    print(' not found')
    sys.exit()



#set up chrome driver and conenct to sites

PATH = "chromedriver.exe"

# creating a driver object that is equal to the chrome driver
driver = webdriver.Chrome(PATH)

# using .get to visit the website
driver.get("https://www.ea.com/games/madden-nfl/player-ratings")


#instructions for what to do on the site




#download list of player names to look for

wb = load_workbook('automated nfl data receiver.xlsx', data_only=True)

ws = wb['Sheet6']
print(ws['A1'].value)
ws['A1']="test"




ovr_dict = {}


#look through rows
for row in range(65):
    row += 1
    a_cell_name = str('A' + str(row))
    if ws[a_cell_name].value:
        print(ws[a_cell_name].value)
        try:
            search = driver.execute_script("return document.querySelector('ea-player-ratings-data-table').shadowRoot.querySelector('#playerRatingsSearch')")
            # time.sleep(3)
            search.clear()
            search.send_keys(ws[a_cell_name].value)
            # search.send_keys('josh allen')
            # print('done')
            search.send_keys(Keys.RETURN)
            time.sleep(1.5)
            ovrvalstep1 = driver.execute_script("return document.querySelector('ea-player-ratings-data-table').shadowRoot.querySelector('#tbody > div:nth-child(1) > div.eapl-player-ratings-data-table__tbody-row-data > div.eapl-player-ratings-data-table__attributes > div:nth-child(4) > span')")
            # print('done')
            # lookingglass.click()
            # print('done')

            # ovrval = driver.find_element_by_xpath('/html/body/main/ea-section[1]/ea-section-column/ea-player-ratings-data-table//div/div[5]/div[1]/div[1]/div[1]/div[4]/span')

            print(ovrvalstep1.get_attribute("innerHTML").splitlines()[0])

            ovrval = ovrvalstep1.get_attribute("innerHTML").splitlines()[0]
            ovr_dict[a_cell_name] = ovrval
            print(ovr_dict)
        except:
            print('no')
            pass

ws2 = wb['Sheet1']


for item in ovr_dict:
    print(item)
    print(ws2[item].value)
    ws2[item] = ovr_dict.get(item)

wb.save('automated nfl data receiver.xlsx')

#update spreadsheet










